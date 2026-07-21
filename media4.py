import os
import re
import html
import urllib.parse
import json
import requests
import feedparser
from datetime import datetime
from dotenv import load_dotenv
import google.generativeai as genai
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Memuat environment variables dari file .env
load_dotenv()

# ------------------------------------------------------------------
# Konfigurasi Dasar & API
# ------------------------------------------------------------------
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
}

# Pemetaan parameter Google News untuk berbagai negara
COUNTRY_MAP = {
    "1": {"hl": "id", "gl": "ID", "ceid": "ID:id", "label": "Indonesia"},
    "2": {"hl": "en", "gl": "US", "ceid": "US:en", "label": "Amerika Serikat (Global / English)"},
    "3": {"hl": "en", "gl": "GB", "ceid": "GB:en", "label": "Inggris (UK)"},
    "4": {"hl": "en", "gl": "SG", "ceid": "SG:en", "label": "Singapura"},
    "5": {"hl": "ms", "gl": "MY", "ceid": "MY:ms", "label": "Malaysia"},
    "6": {"hl": "ja", "gl": "JP", "ceid": "JP:ja", "label": "Jepang"},
    "7": {"hl": "en", "gl": "AU", "ceid": "AU:en", "label": "Australia"}
}

def clean_html(html_text: str) -> str:
    """Menghapus tag HTML dan entitas kode seperti &nbsp; agar teks bersih."""
    if not html_text:
        return ""
    
    clean = html.unescape(html_text)
    clean = clean.replace('&nbsp;', ' ').replace('\xa0', ' ')
    clean = re.sub(r'<[^>]+>', '', clean)
    clean = re.sub(r'\s+', ' ', clean).strip()
    return clean

def extract_image_url(html_text: str) -> str:
    """Mencoba mengambil URL gambar dari HTML summary jika tersedia."""
    if not html_text:
        return "-"
    match = re.search(r'<img[^>]+src=["\']([^"\']+)["\']', html_text)
    if match:
        return match.group(1)
    return "-"

def fetch_news_pure_scraping(topic: str, location: str, time_range_code: str, max_results: int, country_cfg: dict):
    """Mencari berita dari ribuan media resmi terpercaya via Google News RSS berdasarkan konfigurasi negara."""
    
    search_query = f"{topic} {location}".strip()
    
    if time_range_code == "1":
        search_query += " when:1d"  # 24 jam terakhir
    elif time_range_code == "2":
        search_query += " when:7d"  # 7 hari terakhir
    elif time_range_code == "3":
        search_query += " when:30d" # 30 hari terakhir
    
    encoded_query = urllib.parse.quote(search_query)
    
    hl = country_cfg["hl"]
    gl = country_cfg["gl"]
    ceid = country_cfg["ceid"]
    rss_url = f"https://news.google.com/rss/search?q={encoded_query}&hl={hl}&gl={gl}&ceid={ceid}"
    
    all_entries = []
    try:
        print(f"\n[PROSES] Menghubungkan ke server agregator berita negara: {country_cfg['label']}...")
        resp = requests.get(rss_url, headers=HEADERS, timeout=15)
        resp.raise_for_status()
        
        feed = feedparser.parse(resp.content)
        news_items = feed.entries[:max_results] if max_results > 0 else feed.entries

        for entry in news_items:
            title_raw = entry.get("title", "").strip()
            summary_raw = entry.get("summary", "")
            
            source_name = "Media Resmi"
            title_clean = title_raw
            
            if " - " in title_raw:
                parts = title_raw.split(" - ")
                source_name = parts[-1]
                title_clean = " - ".join(parts[:-1])

            title_clean = clean_html(title_clean)
            cleaned_summary = clean_html(summary_raw)
            foto_url = extract_image_url(summary_raw)

            # Menyimpan data teks asli secara utuh agar pemrosesan konteks oleh Gemini lebih akurat
            all_entries.append({
                "raw_title": title_clean,
                "raw_summary": cleaned_summary if cleaned_summary else f"Informasi terkini mengenai {topic}.",
                "Judul": title_clean,          # Placeholder (Akan diganti oleh AI)
                "Sub judul": cleaned_summary,  # Placeholder (Akan diganti oleh AI)
                "Caption": "",                 # Baris baru (Akan diisi oleh AI)
                "Sumber Nama Website": source_name,
                "Foto": foto_url,
                "Sumber": entry.get("link", "")
            })
            
        print(f"[SUKSES] Berhasil mengumpulkan {len(all_entries)} berita resmi mentah.")
    except Exception as e:
        print(f"[GAGAL] Proses scraping berita mengalami masalah: {e}")
        
    return all_entries

def process_content_with_gemini(entries):
    """Memproses judul, sub-judul, dan merancang caption Instagram menggunakan Gemini API."""
    gemini_key = os.getenv("GEMINI_API_KEY")
    if not gemini_key:
        print("[PERINGATAN] GEMINI_API_KEY tidak ditemukan di .env. Menggunakan teks berita asli.")
        return entries

    # Konfigurasi SDK Gemini
    genai.configure(api_key=gemini_key)
    model = genai.GenerativeModel("gemini-1.5-flash")
    
    print(f"\n[PROSES] Memulai pemrosesan AI untuk {len(entries)} berita...")
    
    for idx, item in enumerate(entries, start=1):
        print(f" -> [{idx}/{len(entries)}] Memproses teks AI untuk: {item['raw_title'][:40]}...")
        
        prompt = f"""
        Kamu adalah seorang Senior Social Media Specialist dan Copywriter teknologi terkemuka.
        Tugasmu adalah menyusun artikel berita berikut menjadi materi siap posting untuk konten Instagram (Bahasa Indonesia).

        Informasi Artikel Asli:
        - Judul Asli: {item['raw_title']}
        - Ringkasan Awal: {item['raw_summary']}

        Hasilkan objek JSON dengan struktur key yang tepat seperti di bawah ini:
        1. "judul_instagram": Buat judul yang sangat memikat, bombastis namun edukatif dalam Bahasa Indonesia.
        2. "sub_judul_instagram": Sub judul pendek (maksimal 15 kata) memberikan konteks instan.
        3. "caption_instagram": Buat struktur teks caption lengkap yang terdiri dari:
           - Hook pembuka yang memicu rasa penasaran pembaca.
           - Ringkasan berita utama yang dikemas secara santai dan mudah dipahami orang awam.
           - Penjelasan singkat mengapa berita teknologi ini penting untuk masa depan.
           - Call to Action (CTA) interaktif berupa pertanyaan diskusi untuk meramaikan kolom komentar.
           - Sisipkan 5 hingga 10 hashtag teknologi relevan di baris paling bawah.

        Wajib memberikan respon berupa PURE JSON valid tanpa tambahan teks markdown lain.
        """
        
        try:
            # Memaksa model mengembalikan format JSON terstruktur
            response = model.generate_content(
                prompt,
                generation_config={"response_mime_type": "application/json"}
            )
            
            ai_data = json.loads(response.text)
            
            # Memperbarui entri dengan kreasi dari AI Gemini
            item["Judul"] = ai_data.get("judul_instagram", item["raw_title"])
            item["Sub judul"] = ai_data.get("sub_judul_instagram", item["raw_summary"][:50])
            item["Caption"] = ai_data.get("caption_instagram", "")
            
        except Exception as e:
            print(f"    [GAGAL AI] Gagal memproses artikel ke-{idx}: {e}")
            item["Caption"] = "Gagal memproses caption menggunakan AI."
            
    print("[SUKSES] Seluruh konten berhasil ditransformasi oleh AI Gemini.")
    return entries

def export_to_json(entries, filepath):
    """Menyimpan seluruh berita ke dalam file format JSON terstruktur."""
    # Membersihkan kunci internal mentah sebelum disimpan agar output bersih
    clean_entries = []
    for entry in entries:
        clean_item = {k: v for k, v in entry.items() if not k.startswith('raw_')}
        clean_entries.append(clean_item)

    output_data = {
        "metadata": {
            "total_berita": len(clean_entries),
            "waktu_ekstraksi": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        },
        "berita": clean_entries
    }
    
    try:
        with open(filepath, "w", encoding="utf-8") as json_file:
            json.dump(output_data, json_file, indent=4, ensure_ascii=False)
        print(f"[SUKSES] Data berhasil diexport ke JSON: {filepath}")
    except Exception as e:
        print(f"[GAGAL] Proses export ke file JSON mengalami kendala: {e}")

def export_to_excel(entries, filepath):
    """Menyimpan seluruh berita ke dalam file Excel (.xlsx) dengan styling profesional."""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Hasil Scraping Berita"
        ws.views.sheetView[0].showGridLines = True
        
        # Penambahan kolom Caption ke dalam susunan Header
        headers = ["Judul", "Sub judul", "Caption", "Sumber Nama Website", "Foto", "Sumber"]
        ws.append(headers)
        
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            
        data_font = Font(name="Arial", size=10, bold=False, color="333333")
        zebra_fill = PatternFill(start_color="F2F6F9", end_color="F2F6F9", fill_type="solid")
        
        for row_num, entry in enumerate(entries, 2):
            row_data = [
                entry["Judul"],
                entry["Sub judul"],
                entry["Caption"],
                entry["Sumber Nama Website"],
                entry["Foto"],
                entry["Sumber"]
            ]
            ws.append(row_data)
            
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = data_font
                cell.border = thin_border
                
                if row_num % 2 == 0:
                    cell.fill = zebra_fill
                
                # Pengaturan format per baris kolom
                if col_num in [1, 2, 3]: # Kolom teks panjang
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                elif col_num == 4: # Nama website
                    cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                else: # Link URL
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
                    cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")
        
        ws.row_dimensions[1].height = 28
        for row in range(2, len(entries) + 2):
            ws.row_dimensions[row].height = 65 # Tinggi baris ditambah agar menampung teks caption
            
        # Penyesuaian lebar kolom agar kolom Caption terlihat proporsional
        column_widths = {1: 30, 2: 30, 3: 50, 4: 22, 5: 20, 6: 25}
        for col_idx, width in column_widths.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width
            
        wb.save(filepath)
        print(f"[SUKSES] Data berhasil diexport ke Excel: {filepath}")
    except Exception as e:
        print(f"[GAGAL] Proses export ke file Excel mengalami kendala: {e}")

def main():
    print("=" * 65)
    print("      NEWS AI AUTOMATION BOT FOR INSTAGRAM CONTENT      ")
    print("=" * 65)
    
    # 1. Input Kriteria Pencarian Berita
    topic = input("1. Masukkan Topik Berita (contoh: 'AI', 'Cybersecurity'): ").strip()
    location = input("2. Masukkan Kata Kunci Tambahan/Lokasi (opsional): ").strip()
    
    if not topic:
        print("[EROR] Topik berita minimal harus diisi!")
        return

    # Menu Pilihan Negara Asal Berita
    print("\nPilih Negara Sumber Berita:")
    for key, value in COUNTRY_MAP.items():
        print(f"  [{key}] {value['label']}")
    country_choice = input("Pilih nomor negara (1-7) [Default 1]: ").strip()
    selected_country = COUNTRY_MAP.get(country_choice, COUNTRY_MAP["1"])

    # 2. Input Rentang Waktu Pencarian
    print("\nPilih Rentang Waktu Berita:")
    print("  [1] 24 Jam Terakhir")
    print("  [2] 7 Hari Terakhir")
    print("  [3] 30 Hari Terakhir")
    print("  [4] Semua Waktu (Tanpa Batasan)")
    time_choice = input("Masukkan pilihan rentang waktu (1/2/3/4) [Default 4]: ").strip()
    if time_choice not in ["1", "2", "3", "4"]:
        time_choice = "4"

    # 3. Input Batas Jumlah Berita
    limit_input = input("\nMasukkan jumlah maksimal berita yang ingin diambil [Default 5]: ").strip()
    max_results = int(limit_input) if limit_input.isdigit() else 5

    print("\n" + "-" * 65)
    print(f"[PROSES] Memulai scraping berita tentang '{topic}' untuk wilayah '{selected_country['label']}'...")
    
    # Langkah A: Scraping berita mentah
    entries = fetch_news_pure_scraping(topic, location, time_choice, max_results, selected_country)
    
    if not entries:
        print("\n" + "!" * 65)
        print("[INFO] Tidak ditemukan berita valid dari kriteria tersebut.")
        print("!" * 65)
        return

    # Langkah B: Pemrosesan AI Generatif lewat Gemini API
    entries = process_content_with_gemini(entries)

    print(f"\n[OK] Berhasil mengolah {len(entries)} materi konten berita AI.")
    print("-" * 65)
    
    print("PRATINJAU HASIL BENTUKAN AI GEMINI (3 Teratas):")
    for idx, item in enumerate(entries[:3], start=1):
        print(f" {idx}. 🔥 [JUDUL AI] {item['Judul']}")
        print(f"    📌 [SUB AI] {item['Sub judul']}")
        print(f"    📝 [CAPTION PREVIEW] {item['Caption'][:120]}...\n")

    # Penyusunan nama folder & file output
    folder_name = datetime.now().strftime("%d.%m")
    os.makedirs(folder_name, exist_ok=True)

    clean_filename = re.sub(r'[^a-zA-Z0-9]', '_', f"{topic}_{selected_country['gl']}".strip("_"))
    date_str = datetime.now().strftime("%Y%m%d")
    
    json_filepath = os.path.join(folder_name, f"ai_news_{clean_filename}_{date_str}.json")
    excel_filepath = os.path.join(folder_name, f"ai_news_{clean_filename}_{date_str}.xlsx")
    
    # Langkah C: Ekspor Data Akhir
    export_to_json(entries, json_filepath)
    export_to_excel(entries, excel_filepath)

if __name__ == "__main__":
    main()