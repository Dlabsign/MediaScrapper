# ==================================================================
# CARA MENJALANKAN SCRIPT INI:
# 1. Install library yang dibutuhkan via terminal/cmd:
#    pip install feedparser requests openpyxl
# 2. Jalankan script:
#    python scraper_berita_v2.py
# ==================================================================

import os
import re
import html
import urllib.parse
import json
import requests
import feedparser
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------------
# Konfigurasi Dasar
# ------------------------------------------------------------------
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
}

def clean_html(html_text: str) -> str:
    """Menghapus tag HTML dan entitas kode seperti &nbsp; agar teks bersih."""
    if not html_text:
        return ""
    
    # Konversi entitas HTML (seperti &nbsp; menjadi spasi, &amp; menjadi &)
    clean = html.unescape(html_text)
    
    # Hapus literal teks '&nbsp;' jika masih tertulis manual sebagai string
    clean = clean.replace('&nbsp;', ' ').replace('\xa0', ' ')
    
    # Hapus semua struktur tag HTML (<...>)
    clean = re.sub(r'<[^>]+>', '', clean)
    
    # Bersihkan spasi ganda atau berlebih menjadi satu spasi biasa
    clean = re.sub(r'\s+', ' ', clean).strip()
    return clean

def extract_image_url(html_text: str) -> str:
    """Mencoba mengambil URL gambar dari HTML summary jika tersedia."""
    if not html_text:
        return "-"
    match = re.search(r'<img[^>]+src=["\']([^"\']+)["\']', html_text)
    if match:
        return match.group(1)
    return "-"

def truncate_words(text: str, max_words: int = 15) -> str:
    """Membatasi jumlah kata dalam teks maksimal N kata."""
    words = text.split()
    if len(words) > max_words:
        return " ".join(words[:max_words]) + "..."
    return text

def fetch_news_pure_scraping(topic: str, location: str, time_range_code: str, max_results: int):
    """Mencari berita dari ribuan media resmi terpercaya via Google News RSS."""
    
    search_query = f"{topic} {location}".strip()
    
    if time_range_code == "1":
        search_query += " when:1d"  # 24 jam terakhir
    elif time_range_code == "2":
        search_query += " when:7d"  # 7 hari terakhir
    elif time_range_code == "3":
        search_query += " when:30d" # 30 hari terakhir
    
    encoded_query = urllib.parse.quote(search_query)
    rss_url = f"https://news.google.com/rss/search?q={encoded_query}&hl=id&gl=ID&ceid=ID:id"
    
    all_entries = []
    try:
        print(f"\n[PROSES] Menghubungkan ke server agregator berita...")
        resp = requests.get(rss_url, headers=HEADERS, timeout=15)
        resp.raise_for_status()
        
        feed = feedparser.parse(resp.content)
        news_items = feed.entries[:max_results] if max_results > 0 else feed.entries

        for entry in news_items:
            title_raw = entry.get("title", "").strip()
            summary_raw = entry.get("summary", "")
            
            source_name = "Media Resmi"
            title_clean = title_raw
            
            # Memisahkan judul dengan nama media di bagian akhir ("Judul - Nama Media")
            if " - " in title_raw:
                parts = title_raw.split(" - ")
                source_name = parts[-1]
                title_clean = " - ".join(parts[:-1])

            # Bersihkan teks judul dari kemungkinan entitas HTML
            title_clean = clean_html(title_clean)

            # 1. Judul dibuat ringkas maksimal 15 kata
            judul_ringkas = truncate_words(title_clean, max_words=15)
            
            # 2. Sub judul diambil dan dibersihkan dari potongan summary RSS
            cleaned_summary = clean_html(summary_raw)
            sub_judul = truncate_words(cleaned_summary, max_words=25)
            if not sub_judul or sub_judul == "...":
                sub_judul = f"Informasi terkini mengenai {topic} di {location}."

            # 3. Foto / Gambar pendukung
            foto_url = extract_image_url(summary_raw)

            all_entries.append({
                "Judul": judul_ringkas,
                "Sub judul": sub_judul,
                "Sumber Nama Website": source_name,
                "Foto": foto_url,
                "Sumber": entry.get("link", "")
            })
            
        print(f"[SUKSES] Berhasil mengumpulkan {len(all_entries)} berita resmi.")
    except Exception as e:
        print(f"[GAGAL] Proses scraping berita mengalami masalah: {e}")
        
    return all_entries

def export_to_json(entries, filepath):
    """Menyimpan seluruh berita ke dalam file format JSON terstruktur."""
    output_data = {
        "metadata": {
            "total_berita": len(entries),
            "waktu_ekstraksi": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        },
        "berita": entries
    }
    
    try:
        with open(filepath, "w", encoding="utf-8") as json_file:
            json.dump(output_data, json_file, indent=4, ensure_ascii=False)
        print(f"[SUKSES] Data berhasil diexport ke JSON: {filepath}")
    except Exception as e:
        print(f"[GAGAL] Proses export ke file JSON mengalami kendala: {e}")

def export_to_excel(entries, filepath):
    """Menyimpan seluruh berita ke dalam file Excel (.xlsx) dengan styling profesional."""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Hasil Scraping Berita"
        
        # Aktifkan grid lines
        ws.views.sheetView[0].showGridLines = True
        
        # Header sesuai permintaan user
        headers = ["Judul", "Sub judul", "Sumber Nama Website", "Foto", "Sumber"]
        ws.append(headers)
        
        # Styling Header
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Blue
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            
        # Data Rows Styling
        data_font = Font(name="Arial", size=10, bold=False, color="333333")
        zebra_fill = PatternFill(start_color="F2F6F9", end_color="F2F6F9", fill_type="solid")
        
        for row_num, entry in enumerate(entries, 2):
            row_data = [
                entry["Judul"],
                entry["Sub judul"],
                entry["Sumber Nama Website"],
                entry["Foto"],
                entry["Sumber"]
            ]
            ws.append(row_data)
            
            # Terapkan style ke baris data
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = data_font
                cell.border = thin_border
                
                # Baris genap diberi warna zebra background agar mudah dibaca
                if row_num % 2 == 0:
                    cell.fill = zebra_fill
                
                # Alignment khusus kolom
                if col_num in [1, 2]: # Judul & Subjudul
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                elif col_num == 3: # Nama Website
                    cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                else: # Link Foto & Sumber
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
                    cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")
        
        # Set tinggi baris header & data agar rapi
        ws.row_dimensions[1].height = 28
        for row in range(2, len(entries) + 2):
            ws.row_dimensions[row].height = 40
            
        # Atur lebar kolom secara otomatis dengan batas optimal
        column_widths = {1: 35, 2: 40, 3: 22, 4: 20, 5: 30}
        for col_idx, width in column_widths.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width
            
        wb.save(filepath)
        print(f"[SUKSES] Data berhasil diexport ke Excel: {filepath}")
    except Exception as e:
        print(f"[GAGAL] Proses export ke file Excel mengalami kendala: {e}")

def main():
    print("=" * 65)
    print("        NEWS SCRAPER TO EXCEL & JSON (CLEAN ENTITIES)      ")
    print("=" * 65)
    
    # 1. Input Kriteria Pencarian Berita
    topic = input("1. Masukkan Topik Berita (contoh: 'Gadget', 'Android'): ").strip()
    location = input("2. Masukkan Lokasi (contoh: 'Indonesia', 'Global'): ").strip()
    
    if not topic:
        print("[EROR] Topik berita minimal harus diisi!")
        return

    # 2. Input Rentang Waktu Pencarian
    print("\nPilih Rentang Waktu Berita:")
    print("  [1] 24 Jam Terakhir")
    print("  [2] 7 Hari Terakhir")
    print("  [3] 30 Hari Terakhir")
    print("  [4] Semua Waktu (Tanpa Batasan)")
    time_choice = input("3. Masukkan pilihan rentang waktu (1/2/3/4) [Default 4]: ").strip()
    if time_choice not in ["1", "2", "3", "4"]:
        time_choice = "4"

    # 3. Input Batas Jumlah Berita
    limit_input = input("\n4. Masukkan jumlah maksimal berita yang ingin diambil [Default 15]: ").strip()
    max_results = int(limit_input) if limit_input.isdigit() else 15

    print("\n" + "-" * 65)
    print(f"[PROSES] Memulai scraping berita tentang '{topic}' di '{location}'...")
    
    # Eksekusi Scraping
    entries = fetch_news_pure_scraping(topic, location, time_choice, max_results)
    
    if not entries:
        print("\n" + "!" * 65)
        print("[INFO] Tidak ditemukan berita valid dari situs resmi untuk kriteria tersebut.")
        print("!" * 65)
        return

    print(f"\n[OK] Berhasil mengumpulkan {len(entries)} berita.")
    print("-" * 65)
    
    # Tampilkan 5 berita teratas langsung di terminal sebagai pratinjau
    print("PRATINJAU HASIL BERITA:")
    for idx, item in enumerate(entries[:5], start=1):
        print(f" {idx}. [{item['Sumber Nama Website']}] {item['Judul']}")
        print(f"    Sub: {item['Sub judul']}")
        print(f"    Link: {item['Sumber']}\n")

    # Tentukan Nama Folder Berdasarkan Tanggal (Format DD.MM)
    folder_name = datetime.now().strftime("%d.%m")
    os.makedirs(folder_name, exist_ok=True)

    # Buat Nama File secara dinamis
    clean_filename = re.sub(r'[^a-zA-Z0-9]', '_', f"{topic}_{location}".strip("_"))
    date_str = datetime.now().strftime("%Y%m%d")
    
    json_filepath = os.path.join(folder_name, f"scrape_{clean_filename}_{date_str}.json")
    excel_filepath = os.path.join(folder_name, f"scrape_{clean_filename}_{date_str}.xlsx")
    
    # Ekspor ke file JSON & Excel
    export_to_json(entries, json_filepath)
    export_to_excel(entries, excel_filepath)

if __name__ == "__main__":
    main()