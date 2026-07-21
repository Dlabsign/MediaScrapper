# ==================================================================
# CARA MENJALANKAN SCRIPT INI:
# 1. Install library yang dibutuhkan via terminal/cmd:
#    pip install feedparser requests openpyxl beautifulsoup4
# 2. Jalankan script:
#    python scraper_berita_v3.py
# ==================================================================

import os
import re
import html
import urllib.parse
import json
import requests
import feedparser
from datetime import datetime
from bs4 import BeautifulSoup  # Library baru untuk parsing HTML spesifik
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------------
# Konfigurasi Dasar
# ------------------------------------------------------------------
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
}

def clean_html(html_text: str) -> str:
    """Menghapus tag HTML dan entitas kode seperti &nbsp; agar teks bersih."""
    if not html_text:
        return ""
    clean = html.unescape(html_text)
    clean = clean.replace('&nbsp;', ' ').replace('\xa0', ' ')
    clean = re.sub(r'<[^>]+>', '', clean)
    clean = re.sub(r'\s+', ' ', clean).strip()
    return clean

def extract_image_url(html_text: str) -> str:
    """Mencoba mengambil URL gambar dari HTML summary jika tersedia."""
    if not html_text:
        return "-"
    match = re.search(r'<img[^>]+src=["\']([^"\']+)["\']', html_text)
    if match:
        return match.group(1)
    return "-"

def truncate_words(text: str, max_words: int = 15) -> str:
    """Membatasi jumlah kata dalam teks maksimal N kata."""
    words = text.split()
    if len(words) > max_words:
        return " ".join(words[:max_words]) + "..."
    return text

def fetch_article_description(url: str) -> str:
    """Mengunjungi URL artikel asli untuk mengambil meta description-nya."""
    try:
        # Menghubungi URL (requests otomatis mengikuti redirect dari Google News ke situs asli)
        resp = requests.get(url, headers=HEADERS, timeout=7)
        if resp.status_code == 200:
            soup = BeautifulSoup(resp.content, "html.parser")
            
            # Cari meta tag description standar atau OpenGraph description
            meta_desc = (
                soup.find("meta", attrs={"name": "description"}) or 
                soup.find("meta", attrs={"property": "og:description"}) or
                soup.find("meta", attrs={"name": "twitter:description"})
            )
            
            if meta_desc and meta_desc.get("content"):
                return clean_html(meta_desc.get("content"))
    except Exception:
        # Jika gagal (timeout/blocked), abaikan agar skrip tidak berhenti
        pass
    return "-"

def fetch_news_pure_scraping(topic: str, location: str, time_range_code: str, max_results: int):
    """Mencari berita dari ribuan media resmi terpercaya via Google News RSS + Detail Scraper."""
    
    search_query = f"{topic} {location}".strip()
    if time_range_code == "1":
        search_query += " when:1d"
    elif time_range_code == "2":
        search_query += " when:7d"
    elif time_range_code == "3":
        search_query += " when:30d"
    
    encoded_query = urllib.parse.quote(search_query)
    rss_url = f"https://news.google.com/rss/search?q={encoded_query}&hl=id&gl=ID&ceid=ID:id"
    
    all_entries = []
    try:
        print(f"\n[PROSES] Menghubungkan ke server agregator berita...")
        resp = requests.get(rss_url, headers=HEADERS, timeout=15)
        resp.raise_for_status()
        
        feed = feedparser.parse(resp.content)
        news_items = feed.entries[:max_results] if max_results > 0 else feed.entries

        total_item = len(news_items)
        for idx, entry in enumerate(news_items, 1):
            title_raw = entry.get("title", "").strip()
            summary_raw = entry.get("summary", "")
            article_link = entry.get("link", "")
            
            source_name = "Media Resmi"
            title_clean = title_raw
            
            if " - " in title_raw:
                parts = title_raw.split(" - ")
                source_name = parts[-1]
                title_clean = " - ".join(parts[:-1])

            title_clean = clean_html(title_clean)
            judul_ringkas = truncate_words(title_clean, max_words=15)
            
            cleaned_summary = clean_html(summary_raw)
            sub_judul = truncate_words(cleaned_summary, max_words=25)
            if not sub_judul or sub_judul == "...":
                sub_judul = f"Informasi terkini mengenai {topic} di {location}."

            # --- PROSES BARU: Ambil Deskripsi Asli dari Web Sumber ---
            print(f"[PROGRESS] [{idx}/{total_item}] Mengambil deskripsi artikel dari {source_name}...")
            deskripsi_artikel = fetch_article_description(article_link)
            # Jika deskripsi terlalu panjang, opsional bisa kamu truncate atau biarkan utuh
            deskripsi_artikel = truncate_words(deskripsi_artikel, max_words=50) 
            
            # Coba ambil gambar dari RSS media_content jika ada (Optimasi)
            foto_url = "-"
            if "media_content" in entry and len(entry["media_content"]) > 0:
                foto_url = entry["media_content"][0].get("url", "-")
            if foto_url == "-":
                foto_url = extract_image_url(summary_raw)

            all_entries.append({
                "Judul": judul_ringkas,
                "Sub judul": sub_judul,
                "Deskripsi Artikel": deskripsi_artikel,  # Data Baru
                "Sumber Nama Website": source_name,
                "Foto": foto_url,
                "Sumber": article_link
            })
            
        print(f"\n[SUKSES] Berhasil mengumpulkan {len(all_entries)} berita resmi beserta deskripsinya.")
    except Exception as e:
        print(f"[GAGAL] Proses scraping berita mengalami masalah: {e}")
        
    return all_entries

def export_to_json(entries, filepath):
    """Menyimpan seluruh berita ke dalam file format JSON terstruktur."""
    output_data = {
        "metadata": {
            "total_berita": len(entries),
            "waktu_ekstraksi": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        },
        "berita": entries
    }
    try:
        with open(filepath, "w", encoding="utf-8") as json_file:
            json.dump(output_data, json_file, indent=4, ensure_ascii=False)
        print(f"[SUKSES] Data berhasil diexport ke JSON: {filepath}")
    except Exception as e:
        print(f"[GAGAL] Proses export ke file JSON mengalami kendala: {e}")

def export_to_excel(entries, filepath):
    """Menyimpan seluruh berita ke dalam file Excel (.xlsx) dengan styling profesional."""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Hasil Scraping Berita"
        ws.views.sheetView[0].showGridLines = True
        
        # Ditambahkan kolom "Deskripsi Artikel"
        headers = ["Judul", "Sub judul", "Deskripsi Artikel", "Sumber Nama Website", "Foto", "Sumber"]
        ws.append(headers)
        
        header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
        )
        
        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
            
        data_font = Font(name="Arial", size=10, bold=False, color="333333")
        zebra_fill = PatternFill(start_color="F2F6F9", end_color="F2F6F9", fill_type="solid")
        
        for row_num, entry in enumerate(entries, 2):
            row_data = [
                entry["Judul"],
                entry["Sub judul"],
                entry["Deskripsi Artikel"], # Data Baru
                entry["Sumber Nama Website"],
                entry["Foto"],
                entry["Sumber"]
            ]
            ws.append(row_data)
            
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = data_font
                cell.border = thin_border
                
                if row_num % 2 == 0:
                    cell.fill = zebra_fill
                
                # Alignment & wrapping adjustment
                if col_num in [1, 2, 3]: # Judul, Subjudul, Deskripsi
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                elif col_num == 4: # Nama Website
                    cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
                else: # Link Foto & Sumber
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
                    cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")
        
        ws.row_dimensions[1].height = 28
        for row in range(2, len(entries) + 2):
            ws.row_dimensions[row].height = 55  # Ditinggikan sedikit karena ada deskripsi lebih panjang
            
        # Update lebar kolom secara optimal (Total 6 Kolom)
        column_widths = {1: 30, 2: 35, 3: 45, 4: 20, 5: 15, 6: 25}
        for col_idx, width in column_widths.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width
            
        wb.save(filepath)
        print(f"[SUKSES] Data berhasil diexport ke Excel: {filepath}")
    except Exception as e:
        print(f"[GAGAL] Proses export ke file Excel mengalami kendala: {e}")

def main():
    print("=" * 65)
    print("     NEWS SCRAPER V3 (WITH ARTICLE META-DESCRIPTION)      ")
    print("=" * 65)
    
    topic = input("1. Masukkan Topik Berita (contoh: 'Gadget', 'Android'): ").strip()
    location = input("2. Masukkan Lokasi (contoh: 'Indonesia', 'Global'): ").strip()
    
    if not topic:
        print("[EROR] Topik berita minimal harus diisi!")
        return

    print("\nPilih Rentang Waktu Berita:")
    print("  [1] 24 Jam Terakhir")
    print("  [2] 7 Hari Terakhir")
    print("  [3] 30 Hari Terakhir")
    print("  [4] Semua Waktu (Tanpa Batasan)")
    time_choice = input("3. Masukkan pilihan rentang waktu (1/2/3/4) [Default 4]: ").strip()
    if time_choice not in ["1", "2", "3", "4"]:
        time_choice = "4"

    limit_input = input("\n4. Masukkan jumlah maksimal berita yang ingin diambil [Default 15]: ").strip()
    max_results = int(limit_input) if limit_input.isdigit() else 15
    if max_results <= 0:
        max_results = 15

    print("\n" + "-" * 65)
    print(f"[PROSES] Memulai scraping berita tentang '{topic}' di '{location}'...")
    
    entries = fetch_news_pure_scraping(topic, location, time_choice, max_results)
    
    if not entries:
        print("\n" + "!" * 65)
        print("[INFO] Tidak ditemukan berita valid untuk kriteria tersebut.")
        print("!" * 65)
        return

    folder_name = datetime.now().strftime("%d.%m")
    os.makedirs(folder_name, exist_ok=True)

    clean_filename = re.sub(r'[^a-zA-Z0-9]', '_', f"{topic}_{location}".strip("_"))
    date_str = datetime.now().strftime("%Y%m%d")
    
    json_filepath = os.path.join(folder_name, f"scrape_{clean_filename}_{date_str}.json")
    excel_filepath = os.path.join(folder_name, f"scrape_{clean_filename}_{date_str}.xlsx")
    
    export_to_json(entries, json_filepath)
    export_to_excel(entries, excel_filepath)

if __name__ == "__main__":
    main()